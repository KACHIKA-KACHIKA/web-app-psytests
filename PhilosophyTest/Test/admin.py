from django.http import HttpResponse
from django.contrib import admin
from .models import UserPersonalInfo, Answers, Questions
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from collections import defaultdict
from transliterate import translit
def download_answers_excel(modeladmin, request, queryset):
    workbook = Workbook()
    sheet = workbook.active

    # Заголовки столбцов
    sheet['A1'] = 'ФИО'
    sheet['B1'] = 'Группа'
    themes = [
        ('Тема №1', 1),
        ('Тема №2', 2),
        ('Тема №3', 3),
        ('Тема №4', 4),
        ('Тема №5', 5),
        ('Тема №6', 6),
        ('Тема №7', 7),
        ('Тема №8', 8),
        ('Тема №9', 9),
    ]
    for i, theme in enumerate(themes):
        col_letter = get_column_letter(i + 3)  # Начинаем с третьей колонки
        sheet[f'{col_letter}1'] = theme[0]

    row_num = 2
    for user_info in queryset:
        #answers = user_info.answers_set.all()

        theme_sums = defaultdict(lambda: [0, 0])

        sheet.cell(row=row_num, column=1, value=user_info.user_second_name + user_info.user_name)
        sheet.cell(row=row_num, column=2, value=user_info.user_group)

        for user_answer, id_questions in user_info.answers_list():
            question_number = id_questions.id_question
            theme_index = (question_number - 1) % len(themes)
            theme = themes[theme_index]

            if user_answer.startswith('-'):
                theme_sums[theme[1]][1] += int(user_answer)   # Увеличиваем сумму минусов
            else:
                theme_sums[theme[1]][0] += int(user_answer)    # Увеличиваем сумму плюсов

        for i, theme in enumerate(themes):
            plus_count = theme_sums[theme[1]][0]
            minus_count = theme_sums[theme[1]][1]
            cell_value = f'({plus_count}+)/({abs(minus_count)}-)'
            col_letter = get_column_letter(i + 3)  # Начинаем с третьей колонки
            sheet.cell(row=row_num, column=i + 3, value=cell_value)

    row_num += 1
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    user_second_name_translit = translit(user_info.user_second_name, 'ru', reversed=True)
    excel_name = user_second_name_translit + "_" + user_info.user_group + "_" + "answers"
    response['Content-Disposition'] = f'attachment; filename={excel_name}.xlsx'
    workbook.save(response)
    return response

class AnswersInline(admin.TabularInline):
    model = Answers
    extra = 0

@admin.register(UserPersonalInfo)
class PersonAdmin(admin.ModelAdmin):
    list_display = ['user_second_name', 'user_name', 'user_group', 'id_user', 'answers_list']
    list_display = ['user_second_name', 'user_name', 'user_group', 'answers_list']
    inlines = [AnswersInline]
    actions = [download_answers_excel]